from pathlib import Path
from openpyxl import load_workbook

path = Path('/home/ubuntu/upload/IMCAN-Reference-Sheet---2024(1).xlsm')
terms = ('imcan', 'support', 'no longer', 'not supported', 'cease', 'unsupported', 'supporting')
wb = load_workbook(path, read_only=True, data_only=True, keep_links=False)
for ws in wb.worksheets:
    for row in ws.iter_rows():
        values = [str(cell.value or '') for cell in row]
        joined = ' | '.join(values)
        low = joined.lower()
        if any(term in low for term in terms):
            hits = [f'{cell.coordinate}={cell.value}' for cell in row if cell.value not in (None, '')]
            print(f'[{ws.title}] ' + ' ; '.join(hits))
